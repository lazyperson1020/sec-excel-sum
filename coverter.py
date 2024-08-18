!pip install pandas openpyxl openai
import pandas as pd
from openpyxl import load_workbook
import openai
API_BASE_URL = "https://llama.us.gaianet.network/v1"
MODEL_NAME = "llama"
API_KEY = "GAIA" 
client = openai.OpenAI(base_url=API_BASE_URL, api_key=API_KEY)
def summarize_text(text):
    response = client.chat.completions.create(
        messages=[
            {
                "role": "system",
                "content": "These are the SEC financial reports of a company. Respond with a comprehensive summary of the text given.",
            },
            {
                "role": "user",
                "content": text,
            }
        ],
        model=MODEL_NAME,
        stream=False,
    )
    return response.choices[0].message.content
def generate_summary_for_excel(file_path, summary_output_file):
    wb = load_workbook(file_path)
    all_sentences = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_sentences.append(f"Sheet: {sheet_name}")
        current_section = []
        for row in ws.iter_rows(values_only=False):
            first_col_value = row[0].value
            if row[0].font.bold:
                if current_section:
                    all_sentences.append("\n".join(current_section))
                    current_section = []
                current_section.append(first_col_value)
            else:
                sentence = [f"{row[0].value}:"]
                for idx, cell in enumerate(row[1:], start=1):
                    if cell.value is not None:
                        sentence.append(f"{ws.cell(row=1, column=idx+1).value} - {cell.value}")
                current_section.append(" : ".join(sentence))
        if current_section:
            all_sentences.append("\n".join(current_section))
        all_sentences.append("\n" + "="*40 + "\n")
    temp_text_file = 'temp_output.txt'
    with open(temp_text_file, 'w') as f:
        f.write("\n\n".join(all_sentences))
    with open(temp_text_file, 'r') as f:
        sections = f.read().split("\n" + "="*40 + "\n")  
    summaries = []
    for section in sections:
        if section.strip():  
            summary = summarize_text(section.strip())
            summaries.append(summary)
    with open(summary_output_file, 'w') as f:
        f.write("\n\n".join(summaries))

    print(f"Summaries have been generated and saved to {summary_output_file}")

if __name__ == "__main__":
    input_excel_file = '/content/Financial_Report (1) (1).xlsx'  # Replace with Excel file path
    summary_output_file = 'summary_output.txt'
    generate_summary_for_excel(input_excel_file, summary_output_file)
